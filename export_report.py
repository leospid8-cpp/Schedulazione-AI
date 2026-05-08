import io
from datetime import date
import pandas as pd


def genera_pdf(linea_nome: str, start_day: date, end_day: date,
               df_produzione: pd.DataFrame, eventi: list) -> bytes:
    """Genera PDF di report per una linea nel range di date."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.enums import TA_CENTER

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=2 * cm, bottomMargin=2 * cm)
    styles = getSampleStyleSheet()
    story = []

    ink = colors.HexColor('#1a2538')
    light = colors.HexColor('#eef3ff')

    def _table_style():
        return TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), ink),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, light]),
        ])

    title_style = ParagraphStyle('TitleC', parent=styles['Title'], alignment=TA_CENTER, fontSize=16)
    story.append(Paragraph(f"Report Produzione — {linea_nome}", title_style))
    story.append(Spacer(1, 0.3 * cm))
    story.append(Paragraph(f"Periodo: {start_day} → {end_day}", styles['Normal']))
    story.append(Spacer(1, 0.6 * cm))

    total_ok = int(df_produzione['ok'].sum()) if not df_produzione.empty else 0
    total_ko = int(df_produzione['ko'].sum()) if not df_produzione.empty else 0
    total_target = int(df_produzione['target_ok'].sum()) if not df_produzione.empty else 0
    perc = int(total_ok / total_target * 100) if total_target > 0 else 0

    story.append(Paragraph("Riepilogo", styles['Heading2']))
    summary = [
        ["Metrica", "Valore"],
        ["Pezzi OK", str(total_ok)],
        ["Pezzi KO (scarti)", str(total_ko)],
        ["Target", str(total_target)],
        ["Avanzamento", f"{perc}%"],
    ]
    t = Table(summary, colWidths=[8 * cm, 8 * cm])
    t.setStyle(_table_style())
    story.append(t)
    story.append(Spacer(1, 0.5 * cm))

    story.append(Paragraph("Produzione Giornaliera", styles['Heading2']))
    if not df_produzione.empty:
        rows = [["Giorno", "OK", "KO", "Target"]]
        for _, r in df_produzione.iterrows():
            rows.append([str(r['giorno']), str(int(r['ok'])), str(int(r['ko'])), str(int(r['target_ok']))])
        pt = Table(rows, colWidths=[5 * cm, 3 * cm, 3 * cm, 5 * cm])
        pt.setStyle(_table_style())
        story.append(pt)
    else:
        story.append(Paragraph("Nessun dato nel periodo.", styles['Normal']))
    story.append(Spacer(1, 0.5 * cm))

    fermi = [e for e in eventi if e.get('tipo') in ('START', 'STOP')]
    if fermi:
        story.append(Paragraph("Fermi e Avvii Linea", styles['Heading2']))
        ev_rows = [["Timestamp", "Tipo", "Ordine"]]
        for e in fermi[:50]:
            ev_rows.append([
                str(e.get('ts', ''))[:19],
                str(e.get('tipo', '')),
                str(e.get('ordine_codice') or e.get('order_id') or '—'),
            ])
        et = Table(ev_rows, colWidths=[7 * cm, 3 * cm, 6 * cm])
        et.setStyle(_table_style())
        story.append(et)

    doc.build(story)
    return buf.getvalue()


def genera_excel(linea_nome: str, start_day: date, end_day: date,
                 df_produzione: pd.DataFrame, eventi: list) -> bytes:
    """Genera Excel di report per una linea nel range di date."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    buf = io.BytesIO()
    wb = Workbook()

    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill(fill_type="solid", fgColor="1A2538")
    center = Alignment(horizontal="center")
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    def _hdr(ws):
        for cell in ws[ws.max_row]:
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = center
            cell.border = thin

    def _row(ws):
        for cell in ws[ws.max_row]:
            cell.border = thin
            cell.alignment = center

    total_ok = int(df_produzione['ok'].sum()) if not df_produzione.empty else 0
    total_ko = int(df_produzione['ko'].sum()) if not df_produzione.empty else 0
    total_target = int(df_produzione['target_ok'].sum()) if not df_produzione.empty else 0
    perc = int(total_ok / total_target * 100) if total_target > 0 else 0

    ws = wb.active
    ws.title = "Riepilogo"
    ws['A1'] = f"Report Produzione — {linea_nome}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Periodo: {start_day} → {end_day}"
    ws.append([])
    ws.append(["Metrica", "Valore"])
    _hdr(ws)
    for r in [["Pezzi OK", total_ok], ["Pezzi KO (scarti)", total_ko],
              ["Target", total_target], ["Avanzamento", f"{perc}%"]]:
        ws.append(r)
        _row(ws)
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 15

    ws2 = wb.create_sheet("Produzione Giornaliera")
    ws2.append(["Giorno", "OK", "KO", "Target"])
    _hdr(ws2)
    for _, r in df_produzione.iterrows():
        ws2.append([str(r['giorno']), int(r['ok']), int(r['ko']), int(r['target_ok'])])
        _row(ws2)
    for col in ['A', 'B', 'C', 'D']:
        ws2.column_dimensions[col].width = 14

    ws3 = wb.create_sheet("Eventi")
    ws3.append(["Timestamp", "Tipo", "Ordine", "Qta"])
    _hdr(ws3)
    for e in eventi:
        ws3.append([
            str(e.get('ts', ''))[:19],
            str(e.get('tipo', '')),
            str(e.get('ordine_codice') or e.get('order_id') or '—'),
            int(e.get('qta', 0)),
        ])
        _row(ws3)
    for col in ['A', 'B', 'C', 'D']:
        ws3.column_dimensions[col].width = 20

    wb.save(buf)
    return buf.getvalue()
